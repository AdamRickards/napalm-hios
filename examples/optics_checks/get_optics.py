import re
import json
import ipaddress
from typing import List, Dict, Any, Optional
import logging
import sys
import os
from datetime import datetime
import time
from pathlib import Path
import argparse

# Add this function near the top of your script
def get_resource_path(relative_path: str) -> str:
    """Get absolute path to resource, works for dev and for PyInstaller."""
    if getattr(sys, 'frozen', False):
        # Running in PyInstaller bundle
        return os.path.join(os.path.dirname(sys.executable), relative_path)
    else:
        # Running in normal Python
        return os.path.abspath(relative_path)

def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description='Network Optics Data Collection Tool')
    
    parser.add_argument('-c', '--config', 
                      default='script.cfg',
                      help='Path to configuration file (default: script.cfg)')
    
    parser.add_argument('-o', '--output',
                      default='network_data.xlsx',
                      help='Path to output Excel file (default: network_data.xlsx)')
    
    parser.add_argument('-r', '--retries',
                      type=int,
                      default=3,
                      help='Number of connection retries (default: 3)')
    
    parser.add_argument('-d', '--delay',
                      type=int,
                      default=5,
                      help='Delay between retries in seconds (default: 5)')
    
    parser.add_argument('-t', '--timeout',
                      type=int,
                      default=30,
                      help='Connection timeout in seconds (default: 30)')
    
    parser.add_argument('-v', '--verbose',
                      action='store_true',
                      help='Enable verbose logging')

    return parser.parse_args()

# Rest of the imports and logging setup...
args = parse_arguments()
# Create logs directory in the executable's directory
log_dir = os.path.join(os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.getcwd(), 'logs')
os.makedirs(log_dir, exist_ok=True)
# Update log filename to use the logs directory
log_filename = os.path.join(log_dir, f'optics_collection_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
# Configure logging with timestamp in filename
logging.basicConfig(
    filename=log_filename,
    level=logging.DEBUG if args.verbose else logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Add console handler to show logs in terminal
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.DEBUG if args.verbose else logging.INFO)
formatter = logging.Formatter('%(levelname)s: %(message)s')
console_handler.setFormatter(formatter)
logging.getLogger().addHandler(console_handler)

# Control logging level for external libraries
logging.getLogger('paramiko').setLevel(logging.DEBUG if args.verbose else logging.WARNING)
logging.getLogger('napalm').setLevel(logging.DEBUG if args.verbose else logging.WARNING)
logging.getLogger('netmiko').setLevel(logging.DEBUG if args.verbose else logging.WARNING)

def is_valid_ipv4(ip: str) -> bool:
    """Validate if a string is a valid IPv4 address."""
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        return False

def parse_config(config_file: str) -> tuple[str, str, List[str]]:
    """
    Parse the configuration file for username, password, and IP addresses.
    
    Args:
        config_file: Path to the configuration file
        
    Returns:
        tuple containing username, password, and list of IP addresses
        
    Raises:
        FileNotFoundError: If config file doesn't exist
        ValueError: If config file is invalid or missing required fields
    """
    if not os.path.exists(config_file):
        raise FileNotFoundError(f"Configuration file '{config_file}' not found")
    
    username = ""
    password = ""
    ip_addresses = []
    
    logging.info(f"Reading configuration from {config_file}")
    
    try:
        with open(config_file, 'r') as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                # Skip empty lines and comments
                if not line or line.startswith('#'):
                    continue
                    
                if line.startswith('username'):
                    try:
                        username = line.split(None, 1)[1]
                        logging.debug("Username configuration found")
                    except IndexError:
                        raise ValueError(f"Invalid username format at line {line_num}")
                        
                elif line.startswith('password'):
                    try:
                        password = line.split(None, 1)[1]
                        logging.debug("Password configuration found")
                    except IndexError:
                        raise ValueError(f"Invalid password format at line {line_num}")
                        
                elif is_valid_ipv4(line):
                    ip_addresses.append(line)
                    logging.debug(f"Found valid IP: {line}")
                else:
                    logging.warning(f"Skipping invalid line {line_num}: {line}")
        
        if not username or not password:
            raise ValueError("Configuration file must contain both username and password")
        
        if not ip_addresses:
            raise ValueError("No valid IP addresses found in configuration")
            
        return username, password, ip_addresses
        
    except Exception as e:
        logging.error(f"Error parsing configuration file: {str(e)}")
        raise

def get_device_optics(username: str, password: str, ip_addresses: List[str], 
                     retry_count: int = 3, retry_delay: int = 5,
                     timeout: int = 30) -> Dict[str, Any]:
    """
    Collect optical information from devices using napalm-hios.
    
    Args:
        username: Device login username
        password: Device login password
        ip_addresses: List of device IP addresses to query
        retry_count: Number of connection retries per device
        retry_delay: Delay in seconds between retries
        timeout: Connection timeout in seconds
        
    Returns:
        Dictionary containing optical data for each device
    """
    try:
        logging.debug("Attempting to import napalm_hios...")
        import napalm_hios
        logging.debug(f"napalm_hios loaded from: {napalm_hios.__file__}")
        from napalm import get_network_driver
        driver = get_network_driver('hios')
        logging.debug(f"HIOS driver loaded from: {driver.__module__}")
    except ImportError as e:
        logging.error(f"Import error details: {str(e)}")
        logging.error("Required module napalm-hios not found")
        raise ImportError("napalm-hios module not found. Please install it using: pip install napalm-hios")

    results = {}
    driver = get_network_driver('hios')
    
    for ip in ip_addresses:
        logging.debug(f"Attempting to connect to {ip} using HIOS driver...")
        
        for attempt in range(retry_count):
            try:
                # Connect to device
                device = driver(
                    hostname=ip,
                    username=username,
                    password=password,
                    timeout=timeout
                )
                device.open()
                
                # Get optical information
                logging.debug(f"Connected to {ip}, retrieving optical information...")
                optics_data = device.get_optics()
                results[ip] = optics_data
                
                device.close()
                logging.debug(f"Successfully collected data from {ip}")
                break  # Success - exit retry loop
                
            except Exception as e:
                error_msg = f"Attempt {attempt + 1}/{retry_count} failed for {ip}: {str(e)}"
                logging.error(error_msg)
                
                if attempt < retry_count - 1:
                    logging.info(f"Retrying in {retry_delay} seconds...")
                    time.sleep(retry_delay)
                else:
                    results[ip] = {"error": str(e)}
    
    return results

def export_to_excel(data: Dict[str, Any], output_file: str = 'network_data.xlsx') -> None:
    """
    Convert JSON data to Excel format with specified columns and formula.
    
    Args:
        data: JSON data containing host and interface information
        output_file: Name of the output Excel file
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import PatternFill, Font, Color
        from openpyxl.formatting.rule import FormulaRule
    except ImportError:
        logging.error("Required module openpyxl not found")
        raise ImportError("openpyxl module not found. Please install it using: pip install openpyxl")

    # Ensure output directory exists
    output_dir = os.path.dirname(output_file)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Optics Data {datetime.now().strftime('%Y-%m-%d')}"
    
    # Set up headers with gray background
    headers = ['Host', 'Interface', 'Input Power', 'Output Power', 'Distance', 'Quality', 'Problem']
    header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
    
    # Current row for data insertion
    current_row = 2
    
    # Define data validation for the distance field
    dv = DataValidation(type="decimal", operator="greaterThan", formula1=0)
    dv.error ='Please enter a positive number'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Enter distance in meters'
    dv.promptTitle = 'Distance Input'
    ws.add_data_validation(dv)
    dv.showInputMessage = True
    dv.showErrorMessage = True

    # Process each host and interface
    for host in data:
        if "error" in data[host]:
            # Handle error case
            ws.cell(row=current_row, column=1, value=host)
            ws.cell(row=current_row, column=2, value="ERROR")
            ws.cell(row=current_row, column=3, value=data[host]["error"])
            current_row += 1
            continue

        for interface in data[host]:
            try:
                # Get channel data
                channel_data = data[host][interface]['physical_channels']['channel'][0]['state']
                input_power = channel_data['input_power']['instant']
                output_power = channel_data['output_power']['instant']
                
                # Write data to worksheet
                ws.cell(row=current_row, column=1, value=host)
                ws.cell(row=current_row, column=2, value=interface)
                ws.cell(row=current_row, column=3, value=input_power)
                ws.cell(row=current_row, column=4, value=output_power)
                ws.cell(row=current_row, column=5, value='')
                
                # Add data validation to the Distance cell
                dv.add(f"E{current_row}")
                
                # Add formula for Quality column
                formula = f'=IF(E{current_row}=0,0,C{current_row}/E{current_row})'
                ws.cell(row=current_row, column=6, value=formula)
                
                # Problem detection formula
                formula2 = (
                    f'=IF(OR(F{current_row}=0,COUNTIF(F:F,"<>0")<=1),FALSE,'
                    f'AND('
                    f'  F{current_row}<>0,'
                    f'  ABS((F{current_row}-AVERAGEIF(F:F,"<>0"))/AVERAGEIF(F:F,"<>0"))>0.5'
                    f'))'
                )
                ws.cell(row=current_row, column=7, value=formula2)
                
                current_row += 1
            except (KeyError, IndexError) as e:
                logging.error(f"Error processing data for {host}/{interface}: {str(e)}")
                continue

    # Add conditional formatting for problem rows
    last_col = get_column_letter(len(headers))
    format_range = f'A2:{last_col}1048576'
    red_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
    ws.conditional_formatting.add(
        format_range,
        FormulaRule(
            formula=['=$G2=TRUE'],
            stopIfTrue=True,
            fill=red_fill
        )
    )

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook with error handling
    try:
        wb.save(output_file)
        logging.info(f"Excel report saved successfully to {output_file}")
    except PermissionError:
        alt_filename = f'network_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        logging.warning(f"Could not save to {output_file}, trying alternate filename: {alt_filename}")
        wb.save(alt_filename)
        logging.info(f"Excel report saved successfully to {alt_filename}")

def main():
    """Main function to orchestrate the script execution."""
    start_time = time.time()
    args = parse_arguments()

    try:
        # Check Python version
        if sys.version_info < (3, 7):
            raise RuntimeError("This script requires Python 3.7 or higher")
        
        # Use get_resource_path for config file
        config_path = get_resource_path(args.config)
        # Parse configuration (now using resolved path)
        username, password, ip_addresses = parse_config(config_path)
        # Parse configuration
        logging.info(f"Found {len(ip_addresses)} device(s) to check")
        
        # Collect optical information for all devices
        results = get_device_optics(
            username, password, ip_addresses,
            retry_count=args.retries,
            retry_delay=args.delay,
            timeout=args.timeout
        )
        
        # Export results to Excel
        export_to_excel(results, args.output)
        
        execution_time = time.time() - start_time
        logging.info(f"Script execution completed successfully in {execution_time:.2f} seconds")
        
    except Exception as e:
        logging.error(f"Script execution failed: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
