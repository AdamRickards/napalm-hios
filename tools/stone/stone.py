"""
STONE — SFP Transceiver Optics Network Evaluator

Connects to HiOS switches, reads SFP transceiver optical power levels,
and exports an Excel report with per-port TX/RX power, distance-normalised
quality scores, and automatic outlier detection.

Usage:
    python stone.py
    python stone.py -c my_site.cfg
    python stone.py --protocol snmp
    python stone.py --dry-run
    python stone.py --debug
"""

import sys
import os
import logging
import ipaddress
import argparse
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_resource_path(relative_path: str) -> str:
    """Get absolute path to resource, works for dev and for PyInstaller."""
    if getattr(sys, 'frozen', False):
        return os.path.join(os.path.dirname(sys.executable), relative_path)
    return os.path.abspath(relative_path)


def is_valid_ipv4(s: str) -> bool:
    try:
        ipaddress.ip_address(s)
        return True
    except ValueError:
        return False


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

def parse_arguments():
    parser = argparse.ArgumentParser(
        description='STONE — SFP Transceiver Optics Network Evaluator'
    )
    parser.add_argument('-c', default='script.cfg',
                        help='config file (default: script.cfg)')
    parser.add_argument('-o', default='network_data.xlsx',
                        help='output Excel file (default: network_data.xlsx)')
    parser.add_argument('-r', type=int, default=3,
                        help='connection retries per device (default: 3)')
    parser.add_argument('--delay', type=int, default=5,
                        help='delay between retries in seconds (default: 5)')
    parser.add_argument('-t', type=int, default=30,
                        help='connection timeout in seconds (default: 30)')
    parser.add_argument('--protocol', default=None,
                        choices=['mops', 'snmp', 'ssh'],
                        help='protocol override (default: from config)')
    parser.add_argument('-s', '--silent', action='store_true',
                        help='suppress console output')
    parser.add_argument('--debug', action='store_true',
                        help='verbose protocol logging')
    parser.add_argument('--dry-run', action='store_true',
                        help='show plan without connecting')
    return parser.parse_args()


def parse_config(config_file: str) -> dict:
    """Parse script.cfg into settings and device list."""
    if not os.path.exists(config_file):
        raise FileNotFoundError(f"Configuration file '{config_file}' not found")

    config = {
        'username': '',
        'password': '',
        'protocol': 'mops',
        'devices': [],
    }

    with open(config_file, 'r') as f:
        for line_num, raw_line in enumerate(f, 1):
            line = raw_line.strip()
            if not line or line.startswith('#'):
                continue

            # Key = value pairs
            if '=' in line:
                key, _, val = line.partition('=')
                key = key.strip().lower()
                val = val.strip()

                if key == 'username':
                    config['username'] = val
                elif key == 'password':
                    config['password'] = val
                elif key == 'protocol':
                    config['protocol'] = val.lower()
                else:
                    logging.warning(f"Line {line_num}: unknown setting '{key}'")
                continue

            # Device lines — bare IP
            ip = line.split()[0]
            if is_valid_ipv4(ip):
                config['devices'].append(ip)
            else:
                logging.warning(f"Line {line_num}: skipping invalid IP '{ip}'")

    if not config['username'] or not config['password']:
        raise ValueError("Configuration must contain both username and password")
    if not config['devices']:
        raise ValueError("No valid device IPs found in configuration")

    return config


# ---------------------------------------------------------------------------
# Per-device data gathering (runs in threads)
# ---------------------------------------------------------------------------

def worker_gather(driver, config, ip, timeout, retries, delay):
    """Connect to one device, collect facts + optics data. Retries per-thread."""
    for attempt in range(retries):
        device = None
        try:
            t0 = time.time()

            device = driver(
                hostname=ip,
                username=config['username'],
                password=config['password'],
                timeout=timeout,
                optional_args={'protocol_preference': [config['protocol']]},
            )
            device.open()

            facts = device.get_facts()
            optics = device.get_optics()

            device.close()
            device = None

            # Count SFP ports (interfaces with optics data)
            sfp_count = len(optics)

            dt = time.time() - t0
            return ip, {
                'model': facts.get('model', 'unknown'),
                'hostname': facts.get('hostname', ip),
                'optics': optics,
                'sfp_count': sfp_count,
                'time': dt,
            }, None

        except Exception as e:
            logging.error(f"Attempt {attempt + 1}/{retries} failed for {ip}: {e}")
            if device:
                try:
                    device.close()
                except Exception:
                    pass
            if attempt < retries - 1:
                time.sleep(delay)
            else:
                return ip, None, str(e)

    return ip, None, "all retries exhausted"


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_to_excel(data: dict, output_file: str) -> str:
    """Export optics data to Excel with formulas and conditional formatting.

    Returns the actual filename used (may differ if PermissionError).
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import PatternFill
        from openpyxl.formatting.rule import FormulaRule
    except ImportError:
        raise ImportError("openpyxl not found — pip install openpyxl")

    output_dir = os.path.dirname(output_file)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Optics Data {datetime.now().strftime('%Y-%m-%d')}"

    # Headers
    headers = ['Host', 'Interface', 'Input Power', 'Output Power',
               'Distance', 'Quality', 'Problem']
    header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE",
                              fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill

    current_row = 2

    # Distance field validation
    dv = DataValidation(type="decimal", operator="greaterThan", formula1=0)
    dv.error = 'Please enter a positive number'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Enter distance in meters'
    dv.promptTitle = 'Distance Input'
    dv.showInputMessage = True
    dv.showErrorMessage = True
    ws.add_data_validation(dv)

    for host, result in data.items():
        if 'error' in result:
            ws.cell(row=current_row, column=1, value=host)
            ws.cell(row=current_row, column=2, value="ERROR")
            ws.cell(row=current_row, column=3, value=result['error'])
            current_row += 1
            continue

        optics = result['optics']
        for interface in optics:
            try:
                channel = optics[interface]['physical_channels']['channel'][0]['state']
                input_power = channel['input_power']['instant']
                output_power = channel['output_power']['instant']

                ws.cell(row=current_row, column=1, value=host)
                ws.cell(row=current_row, column=2, value=interface)
                ws.cell(row=current_row, column=3, value=input_power)
                ws.cell(row=current_row, column=4, value=output_power)
                ws.cell(row=current_row, column=5, value='')

                dv.add(f"E{current_row}")

                # Quality = Input Power / Distance
                ws.cell(row=current_row, column=6,
                        value=f'=IF(E{current_row}=0,0,C{current_row}/E{current_row})')

                # Problem = outlier detection (>50% deviation from average)
                ws.cell(row=current_row, column=7,
                        value=(
                            f'=IF(OR(F{current_row}=0,COUNTIF(F:F,"<>0")<=1),FALSE,'
                            f'AND('
                            f'  F{current_row}<>0,'
                            f'  ABS((F{current_row}-AVERAGEIF(F:F,"<>0"))/AVERAGEIF(F:F,"<>0"))>0.5'
                            f'))'
                        ))

                current_row += 1
            except (KeyError, IndexError) as e:
                logging.error(f"Error processing {host}/{interface}: {e}")
                continue

    # Conditional formatting — red fill for problem rows
    last_col = get_column_letter(len(headers))
    red_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0',
                           fill_type='solid')
    ws.conditional_formatting.add(
        f'A2:{last_col}1048576',
        FormulaRule(formula=['=$G2=TRUE'], stopIfTrue=True, fill=red_fill)
    )

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Save with PermissionError fallback
    try:
        wb.save(output_file)
        return output_file
    except PermissionError:
        alt = f'network_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        logging.warning(f"Cannot save to {output_file}, using {alt}")
        wb.save(alt)
        return alt


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = parse_arguments()

    # Silent mode
    if args.silent:
        sys.stdout = open(os.devnull, 'w')

    # Logging setup
    log_dir = os.path.join(
        os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.getcwd(),
        'logs'
    )
    os.makedirs(log_dir, exist_ok=True)
    log_filename = os.path.join(log_dir, f'stone_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')

    log_level = logging.DEBUG if args.debug else logging.INFO
    logging.basicConfig(
        filename=log_filename,
        level=log_level,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

    console = logging.StreamHandler()
    console.setLevel(logging.DEBUG if args.debug else logging.WARNING)
    console.setFormatter(logging.Formatter('%(levelname)s: %(message)s'))
    logging.getLogger().addHandler(console)

    lib_level = logging.DEBUG if args.debug else logging.WARNING
    for lib in ('paramiko', 'napalm', 'netmiko', 'urllib3', 'requests'):
        logging.getLogger(lib).setLevel(lib_level)
    if args.debug:
        logging.getLogger('napalm_hios.mops_client').setLevel(logging.DEBUG)

    start_time = time.time()

    try:
        config_path = get_resource_path(args.c)
        config = parse_config(config_path)

        # CLI override
        if args.protocol:
            config['protocol'] = args.protocol

        # --- Banner ---
        print("\n" + "=" * 60)
        print("  STONE \u2014 SFP Transceiver Optics Network Evaluator")
        print("=" * 60)
        print(f"  Protocol:  {config['protocol'].upper()}"
              f" | Devices: {len(config['devices'])}"
              f" | Retries: {args.r}")
        print("-" * 60)

        if args.dry_run:
            print("\n  Devices:")
            for ip in config['devices']:
                print(f"    {ip}")
            print(f"\n  Output: {args.o}")
            print("\n  [DRY RUN] No connections will be made.\n")
            return

        from napalm import get_network_driver
        driver = get_network_driver('hios')

        # --- Gather optics data in parallel ---
        device_data = {}
        failures = []
        total_sfps = 0

        with ThreadPoolExecutor(max_workers=len(config['devices'])) as pool:
            futures = {
                pool.submit(worker_gather, driver, config, ip,
                            args.t, args.r, args.delay): ip
                for ip in config['devices']
            }
            for future in as_completed(futures):
                ip, data, err = future.result()
                if data:
                    device_data[ip] = data
                    total_sfps += data['sfp_count']
                    sfp_label = 'SFP port' if data['sfp_count'] == 1 else 'SFP ports'
                    print(f"  [{data['time']:5.1f}s] {ip:<17s}"
                          f"{data['model']:<25s}"
                          f"{data['sfp_count']} {sfp_label}")
                else:
                    failures.append((ip, err))
                    print(f"  [FAIL ] {ip:<17s}{err}")

        if not device_data:
            print("\n  FATAL: No devices reachable.\n")
            sys.exit(1)

        # --- Build export data (keyed by IP, with optics or error) ---
        export_data = {}
        for ip in config['devices']:
            if ip in device_data:
                export_data[ip] = device_data[ip]
            else:
                err = next((e for fip, e in failures if fip == ip), 'unknown error')
                export_data[ip] = {'error': err}

        actual_file = export_to_excel(export_data, args.o)

        # --- Footer ---
        elapsed = time.time() - start_time
        print("\n" + "=" * 60)
        print(f"  {len(device_data)}/{len(config['devices'])} devices reached"
              f" | {total_sfps} SFP ports"
              f" | Done in {elapsed:.1f}s")
        print(f"  Output: {actual_file}")
        print("=" * 60 + "\n")

    except Exception as e:
        logging.error(f"Fatal error: {e}")
        print(f"\n  FATAL: {e}\n")
        sys.exit(1)


if __name__ == "__main__":
    main()
